# ------------------- bibliotecas -------------------
from openpyxl import load_workbook
from datetime import datetime
import os
from class_manager import constantes, funções

c = constantes()
f = funções()

# ------------------- definições de planilha ---------------------
wb = load_workbook(filename=c.rota_sample)
ws = wb['status_transf']

# ------------------ execução ----------------------
f.limpar_pasta(c.rota_erros)

num_linha = 2
ult_linha = len(os.listdir(c.rota_input)) + 1
num_arq = 0

if num_linha > ult_linha:

    print('pasta de input vazia')

else:

    while num_linha <= ult_linha:

        pdf = os.listdir(c.rota_input)[num_arq] # <-- nome do arquivo pdf
        cnpj = pdf.split('_')[1].split('.')[0] # <-- numero do cnpj
        rota_pdf = os.path.join(c.rota_input, pdf) # <-- rota do pdf no input
        rota_fundo = os.path.join(c.rota_fundos, cnpj) # <-- rota da pasta do fundo
        cnpj_pnl = ws.cell(row=num_linha, column=1) # <-- celula do cnpj
        status_transf = ws.cell(row=num_linha, column=2) # <-- celula do status do download

        if os.path.exists(rota_fundo):
            f.pasta_existente(rota_fundo, cnpj_pnl, cnpj, status_transf, rota_pdf, pdf)
        else:
            f.pasta_inexistente(cnpj_pnl, cnpj, status_transf, c.rota_erros, pdf, rota_pdf)

        num_linha += 1
        num_arq += 1

# ------------------ Registro das transferências ------------------
wb.save(filename=f'{c.rota_transf_log}/registro_transf_{datetime.now().strftime('%H_%M_%S.%d_%m_%Y')}.xlsx')


