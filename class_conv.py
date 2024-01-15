import sys
from openpyxl import load_workbook

class constants():

    def __init__(self):
        self.rota_planilha = r'rota'
        self.rota_portal_conv = r'rota'
        self.rota_conv_ass = r'rota'

        wb = load_workbook(filename=self.rota_planilha)
        self.mascara = wb['Máscara']
        self.conv = wb['Convocações']
        self.linhas = len([linha for linha in self.conv if not all([celula.value == None for celula in linha])])

        self.nome_AGO = self.mascara.cell(row=9, column=4)
        self.fy_boot = self.mascara.cell(row=9, column=6)
        self.dt_inicio = self.mascara.cell(row=9, column=8)
        self.hr_inicio = self.mascara.cell(row=9, column=10)
        self.dt_fim = self.mascara.cell(row=9, column=12)
        self.hr_fim = self.mascara.cell(row=9, column=14)
        self.dt_presencial = self.mascara.cell(row=11, column=4)
        self.hr_presencial = self.mascara.cell(row=11, column=6)
        self.dt_termino = self.mascara.cell(row=11, column=8)
        self.hr_termino = self.mascara.cell(row=11, column=10)
        self.aprov_automatica = self.mascara.cell(row=11, column=14)
        self.preambulo = self.mascara.cell(row=14, column=3)
        self.deliberacao = self.mascara.cell(row=16, column=3)

        self.nome_AGO_xpath = r'xpath'
        self.fy_boot_xpath = r'xpath'
        self.dt_inicio_xpath = r'xpath'
        self.dt_fim_xpath = r'xpath'
        self.dt_presencial_xpath = r'xpath'
        self.dt_termino_xpath = r'xpath'
        self.tp_evento_xpath = r'xpath'
        self.tp_presenca_xpath = r'xpath'
        self.responsavel_xpath = r'xpath'
        self.analista_xpath = r'xpath'
        self.coordenador_xpath = r'xpath'
        self.aprov_automatica_xpath = r'xpath'
        self.reprov_automatica_xpath = r'xpath'
        self.preambulo_xpath = r'xpath'
        self.deliberacao_xpath = r'xpath'
        self.ins_lote_xpath = r'xpath'
        self.frmWork_lotes_xpath = r'xpath'
        self.ins_conv_xpath = r'path'
        self.conv_ass_xpath = r'path'

class functions():

    def clicar(self, webdriver, xpath):
        webdriver.find_element('xpath', xpath).click()

    def escrever(self, webdriver, xpath, info):
        webdriver.find_element('xpath', xpath).send_keys(info)
    
    def carregar_link(self, webdriver, url):
        webdriver.get(url)

    def verificar_login(self):
        vld_login = input("insira 's' se logado e 'n' caso não: ")
        if vld_login == 's':
            print('logado')
        else:
            print('Para efetuar convocações, é preciso estar logado.')
            sys.exit()

    def verificar_aprovacao(self, chrome, aprov_xpath, reprov_xpath, valor_aprov):
        if valor_aprov == 1:
            self.clicar(chrome, aprov_xpath)
        elif valor_aprov == 2:
            self.clicar(chrome, reprov_xpath)
        else:
            print('insira uma valor válido para aprovação automática na planilha')
            sys.exit
        
