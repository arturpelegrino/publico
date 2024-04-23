from recursos import rotas, xpaths


from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from openpyxl import load_workbook
import os, time

xpath = xpaths()
rota = rotas()

wb = load_workbook(filename=rota.dados)
ws = wb['alterações']
linhas = len([linha for linha in ws if not all([celula.value == None for celula in linha])])

os.system("start chrome.exe --remote-debugging-port=9223")

input('Você está na tela do buscador? ')

driverOptions = Options()
driverOptions.add_experimental_option("debuggerAddress", "127.0.0.1:9223")

driver = webdriver.Chrome(options=driverOptions)

driver.switch_to.window(driver.window_handles[0])

def verificar_carregamento(driver) -> None:
    while driver.find_element('xpath', xpath.validador).get_attribute('style') != 'display: none;':
        time.sleep(1.5)


for i in range(2, linhas + 1):

    cnpj_fundo = ws.cell(row=i, column=1).value
    cnpj_auditor = ws.cell(row=i, column=2).value
    data_alteracao = ws.cell(row=i, column=3).value

    comandos = [['send_keys', xpath.cnpj_fundo, cnpj_fundo.replace('.', '').replace('-','').replace('/', '')],
                ['click', xpath.buscar_fundo],
                ['click', xpath.participantes],
                ['click', xpath.alterar_auditor],
                ['send_keys', xpath.cnpj_auditor, cnpj_auditor.replace('.', '').replace('-','').replace('/', '')],
                ['click', xpath.data_alteracao],
                ['send_keys', xpath.data_alteracao, data_alteracao.replace('/', '')],
                ['click_cancelar', xpath.cancelar_alteracao],
                ['click', xpath.voltar_perfil],
                ['click', xpath.voltar_buscador]]

    for requisitos in comandos:
        if requisitos[0] == 'click':
            driver.find_element('xpath', requisitos[1]).click()
            verificar_carregamento(driver)
        elif requisitos[0] == 'send_keys':
            driver.find_element('xpath', requisitos[1]).send_keys(requisitos[2])
            verificar_carregamento(driver)
        elif requisitos[0] == 'click_cancelar':
            driver.find_element('xpath', requisitos[1]).click()
            verificar_carregamento(driver)
            status = ws.cell(row=i, column=4)
            status.value = 'alterado'
            wb.save(filename=rota.dados)