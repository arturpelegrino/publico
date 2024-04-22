from recursos import rotas, xpaths


from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from openpyxl import load_workbook
import os, time

xpath = xpaths()
rota = rotas()

wb = load_workbook(filename=rota.dados)
ws = wb['alterações']
linhas = len([linha for linha in ws if not all([celula.value == None for celula in linha])])

os.system("start chrome.exe --remote-debugging-port=9223")

input('Você está na tela do buscador? ')

driverOptions = Options()
driverOptions.add_experimental_option("debuggerAddress", "127.0.0.1:9223")

driver = webdriver.Chrome(options=driverOptions)

driver.switch_to.window(driver.window_handles[0])

i = 2

for i in range(2, linhas + 1):

    time.sleep(1)

    cnpj_fundo = ws.cell(row=i, column=1).value
    driver.find_element('xpath', xpath.cnpj_fundo).send_keys(cnpj_fundo)

    time.sleep(1)

    driver.find_element('xpath', xpath.buscar_fundo).click()

    time.sleep(3)

    driver.find_element('xpath', xpath.participantes).click()

    time.sleep(1)

    driver.find_element('xpath', xpath.alterar_auditor).click()

    time.sleep(1)

    cnpj_auditor = ws.cell(row=i, column=2).value
    driver.find_element('xpath', xpath.cnpj_auditor).send_keys(cnpj_auditor.replace('.', '').replace('-', '').replace('/', ''))
    data_alteracao = ws.cell(row=i, column=3).value
    driver.find_element('xpath', xpath.data_alteracao).click()
    driver.find_element('xpath', xpath.data_alteracao).send_keys(data_alteracao.replace('/', ''))

    time.sleep(1)

    driver.find_element('xpath', xpath.cancelar_alteracao).click()

    status = ws.cell(row=i, column=4)
    status.value = 'alterado'
    wb.save(filename=rota.dados)

    time.sleep(1)

    driver.find_element('xpath', xpath.voltar_perfil).click()

    time.sleep(3)

    driver.find_element('xpath', xpath.voltar_buscador).click()

    time.sleep(1)