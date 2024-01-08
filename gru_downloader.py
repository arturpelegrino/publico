from selenium import webdriver
from openpyxl import load_workbook
from datetime import datetime, timedelta
import time, sys
from class_downloader import constantes, funções

c = constantes()
f = funções()

# --------------------- Planiha e afins --------------------------

wb = load_workbook(filename=c.rota_planilha)
ws = wb['fundos']

rows = len([row for row in ws if not all([cell.value == None for cell in row])]) # <-- conta o número de linhas não vazias

i = 2 # <-- linha em que o 1° CNPJ aparece

cpf_obj = ws.cell(row=2, column=10)
cpf = cpf_obj.value

if cpf == None:
    print('insira cpf na planillha')
    sys.exit()

# --------------------- Setup do driver -------------------

chrome_options = webdriver.ChromeOptions()                  ######################################################
prefs = {"download.default_directory" : c.rota_validador}     ##### Define a rota do coletor como pasta de download
chrome_options.add_experimental_option('prefs', prefs)      ######################################################

chrome = webdriver.Chrome(options=chrome_options)           # <-- inicia chrome

# --------------------- Execução --------------------------

f.limpar_pasta(c.rota_coletor)

while i <= rows:

    cnpj = ws.cell(row=i, column=1).value
    status_download = ws.cell(row=i, column=7)
    agora = datetime.now()
    depois = agora + timedelta(seconds=180)
    sair = False

    chrome.get(c.rota_cvm(cnpj, cpf))

    while agora <= depois and not sair:

        f.limpar_pasta(c.rota_validador)

        try:
            f.clicar(chrome, xpath='//*[@id="rblTabela_1"]')  # <-- seleciona o tipo de taxa
            f.clicar(chrome, xpath='//*[@id="ddlTpReg"]/option[30]')  # <-- seleciona o tipo de registro
            f.digitar(chrome, xpath='//*[@id="tbEndereco"]', info=ws.cell(row=i, column=3).value) # <-- escreve endereço
            f.digitar(chrome, xpath='//*[@id="tbCidade"]', info=ws.cell(row=i, column=4).value) # <-- escreve cidade
            f.digitar(chrome, xpath='//*[@id="tbUF"]', info=ws.cell(row=i, column=5).value) # <-- escreve UF
            f.digitar(chrome, xpath='//*[@id="tbCEP"]', info=ws.cell(row=i, column=6).value) # <-- escreve CEP
            f.clicar(chrome, xpath='//*[@id="btnImpGRU"]')  # <-- seleciona botão de download
            time.sleep(5)
            sair = f.processar_download(c.rota_validador, c.rota_coletor, status_download, cnpj)
        except:
            status_download.value = "cnpj inválido"
            sair = True

        agora = datetime.now()

    i += 1

chrome.quit()

wb.save(filename=f'{c.rota_gru_log}/registro_download_{datetime.now().strftime('%H_%M_%S.%d_%m_%Y')}.xlsx')

f.limpar_input(wb, ws, rows, cpf_obj, c.rota_planilha)